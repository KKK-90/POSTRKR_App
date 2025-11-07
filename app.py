# app.py  -- complete file
from flask import Flask, jsonify, request, send_from_directory, render_template, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from sqlalchemy import func
import os
import json
from datetime import datetime
from io import BytesIO
import tempfile
import traceback

# Optional: for xlsx import/export
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_PATH = os.path.join(DATA_DIR, 'postrkr.db')
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app, supports_credentials=True)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DB_PATH
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --------------------
# DB model
# --------------------
class Location(db.Model):
    __tablename__ = 'locations'
    id = db.Column(db.Integer, primary_key=True)
    slNo = db.Column(db.Integer, nullable=True)
    division = db.Column(db.String, nullable=True)
    postOfficeName = db.Column(db.String, nullable=True)
    postOfficeId = db.Column(db.String, nullable=True)
    officeType = db.Column(db.String, nullable=True)
    contactPersonName = db.Column(db.String, nullable=True)
    contactPersonNo = db.Column(db.String, nullable=True)
    altContactNo = db.Column(db.String, nullable=True)
    contactEmail = db.Column(db.String, nullable=True)
    locationAddress = db.Column(db.String, nullable=True)
    location = db.Column(db.String, nullable=True)
    city = db.Column(db.String, nullable=True)
    state = db.Column(db.String, nullable=True)
    pincode = db.Column(db.String, nullable=True)
    numberOfPosToBeDeployed = db.Column(db.Integer, default=0)
    typeOfPosTerminal = db.Column(db.String, nullable=True)
    dateOfReceiptOfDevice = db.Column(db.String, nullable=True)
    noOfDevicesReceived = db.Column(db.Integer, default=0)
    serialNo = db.Column(db.String, nullable=True)
    installationStatus = db.Column(db.String, nullable=True)
    functionalityStatus = db.Column(db.String, nullable=True)
    issuesIfAny = db.Column(db.String, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def as_dict(self):
        return {c.name: getattr(self, c.name) for c in self.__table__.columns}

# create DB
with app.app_context():
    db.create_all()

# --------------------
# Helper
# --------------------
def next_slno():
    max_sl = db.session.query(func.max(Location.slNo)).scalar()
    return (max_sl or 0) + 1

# --------------------
# Routes - Page serve
# --------------------
@app.route('/')
def index():
    # serve template (copy your current index.html into templates/index.html)
    return render_template('index.html')

# static files are served automatically from /static

# --------------------
# API endpoints
# --------------------
@app.route('/api/login', methods=['POST'])
def api_login():
    body = request.get_json() or {}
    username = body.get('username')
    # Simple username check: keep same users list as front-end
    allowed = ["KARNA", "NKR", "SKR", "BGR", "SBI_DOP"]
    if username in allowed:
        # set a simple session cookie (optionally implement Flask-Login later)
        resp = jsonify({"ok": True, "username": username})
        resp.set_cookie('postrkr_user', username, httponly=True)
        return resp
    return jsonify({'ok': False, 'error': 'Unknown user'}), 401

@app.route('/api/logout', methods=['POST'])
def api_logout():
    resp = jsonify({"ok": True})
    resp.delete_cookie('postrkr_user')
    return resp

# list locations
@app.route('/api/locations', methods=['GET'])
def list_locations():
    try:
        rows = Location.query.order_by(Location.slNo.asc()).all()
        return jsonify([r.as_dict() for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# create new location
@app.route('/api/locations', methods=['POST'])
def create_location():
    try:
        payload = request.get_json()
        loc = Location(
            slNo = payload.get('slNo') or next_slno(),
            division=payload.get('division'),
            postOfficeName=payload.get('postOfficeName'),
            postOfficeId=payload.get('postOfficeId'),
            officeType=payload.get('officeType'),
            contactPersonName=payload.get('contactPersonName'),
            contactPersonNo=payload.get('contactPersonNo'),
            altContactNo=payload.get('altContactNo'),
            contactEmail=payload.get('contactEmail'),
            locationAddress=payload.get('locationAddress'),
            location=payload.get('location'),
            city=payload.get('city'),
            state=payload.get('state'),
            pincode=payload.get('pincode'),
            numberOfPosToBeDeployed=payload.get('numberOfPosToBeDeployed') or 0,
            typeOfPosTerminal=payload.get('typeOfPosTerminal'),
            dateOfReceiptOfDevice=payload.get('dateOfReceiptOfDevice'),
            noOfDevicesReceived=payload.get('noOfDevicesReceived') or 0,
            serialNo=payload.get('serialNo'),
            installationStatus=payload.get('installationStatus'),
            functionalityStatus=payload.get('functionalityStatus'),
            issuesIfAny=payload.get('issuesIfAny') or 'None'
        )
        db.session.add(loc)
        db.session.commit()
        return jsonify(loc.as_dict()), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# update
@app.route('/api/locations/<int:id>', methods=['PUT'])
def update_location(id):
    try:
        loc = Location.query.get_or_404(id)
        payload = request.get_json() or {}
        for k, v in payload.items():
            if hasattr(loc, k):
                setattr(loc, k, v)
        db.session.commit()
        return jsonify(loc.as_dict())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# delete
@app.route('/api/locations/<int:id>', methods=['DELETE'])
def delete_location(id):
    try:
        loc = Location.query.get_or_404(id)
        db.session.delete(loc)
        db.session.commit()
        # reassign slNo sequentially
        rows = Location.query.order_by(Location.id).all()
        for idx, r in enumerate(rows, start=1):
            r.slNo = idx
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# import XLSX - server-side (optional)
@app.route('/api/import', methods=['POST'])
def api_import():
    """
    Accepts an uploaded Excel file (form-data key 'file') and replaces DB with its contents.
    Excel should follow the template columns similar to the front-end.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file'}), 400
        f = request.files['file']
        wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.values)
        if len(rows) < 2:
            return jsonify({'error': 'No data rows found'}), 400

        # Clear existing
        Location.query.delete()
        db.session.commit()

        next_id = 1
        for i, row in enumerate(rows[1:], start=1):
            # mapping based on template in your index.html (columns positions)
            # safe access: use index positions from template: see index.html template header.
            # For robustness, read by expected indices
            try:
                division = row[1] or ''
                postOfficeName = row[2] or ''
                postOfficeId = row[3] or f"AUTO-{datetime.utcnow().timestamp()}"
                city = row[11] or ''
                state = row[12] or ''
                numberOfPos = int(row[14]) if row[14] else 0
            except Exception:
                division = postOfficeName = postOfficeId = city = state = ''
                numberOfPos = 0

            loc = Location(
                slNo = i,
                division=division,
                postOfficeName=postOfficeName,
                postOfficeId=postOfficeId,
                city=city,
                state=state,
                numberOfPosToBeDeployed=numberOfPos,
                installationStatus=(row[19] or 'Pending') if len(row) > 19 else 'Pending',
                functionalityStatus=(row[20] or 'Not Tested') if len(row) > 20 else 'Not Tested',
                issuesIfAny=(row[21] or 'None') if len(row) > 21 else 'None'
            )
            db.session.add(loc)
        db.session.commit()
        return jsonify({'ok': True, 'imported': Location.query.count()})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# export to excel
@app.route('/api/export', methods=['GET'])
def api_export():
    try:
        rows = Location.query.order_by(Location.slNo.asc()).all()
        wb = Workbook()
        ws = wb.active
        header = [
            'Sl.No.', 'Division', 'POST OFFICE NAME', 'Post Office ID', 'Office Type',
            'NAME OF CONTACT PERSON AT THE LOCATION', 'CONTACT PERSON NO.', 'ALT CONTACT PERSON NO.',
            'CONTACT EMAIL ID', 'LOCATION ADDRESS', 'LOCATION', 'CITY', 'STATE', 'PINCODE',
            'NUMBER OF POS TO BE_DEPLOYED', 'TYPE OF POS TERMINAL', 'Date of receipt of device',
            'No of devices received', 'Serial No', 'Installation status',
            'Functionality / Working status of POS machines', 'Issues if any'
        ]
        ws.append(header)
        for r in rows:
            ws.append([
                r.slNo, r.division, r.postOfficeName, r.postOfficeId, r.officeType,
                r.contactPersonName, r.contactPersonNo, r.altContactNo, r.contactEmail,
                r.locationAddress, r.location, r.city, r.state, r.pincode,
                r.numberOfPosToBeDeployed, r.typeOfPosTerminal, r.dateOfReceiptOfDevice,
                r.noOfDevicesReceived, r.serialNo, r.installationStatus,
                r.functionalityStatus, r.issuesIfAny
            ])
        tmp = BytesIO()
        wb.save(tmp)
        tmp.seek(0)
        filename = f"POS_Data_Export_{datetime.utcnow().date().isoformat()}.xlsx"
        return send_file(tmp, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# backup
@app.route('/api/backup', methods=['GET'])
def api_backup():
    try:
        rows = [r.as_dict() for r in Location.query.order_by(Location.slNo.asc()).all()]
        payload = {
            'locations': rows,
            'backupDate': datetime.utcnow().isoformat()
        }
        tmp = BytesIO()
        tmp.write(json.dumps(payload, indent=2).encode('utf-8'))
        tmp.seek(0)
        return send_file(tmp, as_attachment=True, download_name=f"POS_Backup_{datetime.utcnow().date().isoformat()}.json", mimetype='application/json')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# restore
@app.route('/api/restore', methods=['POST'])
def api_restore():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        payload = json.load(f)
        rows = payload.get('locations', [])
        Location.query.delete()
        db.session.commit()
        for i, item in enumerate(rows, start=1):
            loc = Location(
                slNo = item.get('slNo') or i,
                division=item.get('division'),
                postOfficeName=item.get('postOfficeName'),
                postOfficeId=item.get('postOfficeId'),
                officeType=item.get('officeType'),
                contactPersonName=item.get('contactPersonName'),
                contactPersonNo=item.get('contactPersonNo'),
                altContactNo=item.get('altContactNo'),
                contactEmail=item.get('contactEmail'),
                locationAddress=item.get('locationAddress'),
                location=item.get('location'),
                city=item.get('city'),
                state=item.get('state'),
                pincode=item.get('pincode'),
                numberOfPosToBeDeployed=item.get('numberOfPosToBeDeployed') or 0,
                typeOfPosTerminal=item.get('typeOfPosTerminal'),
                dateOfReceiptOfDevice=item.get('dateOfReceiptOfDevice'),
                noOfDevicesReceived=item.get('noOfDevicesReceived') or 0,
                serialNo=item.get('serialNo'),
                installationStatus=item.get('installationStatus'),
                functionalityStatus=item.get('functionalityStatus'),
                issuesIfAny=item.get('issuesIfAny') or 'None'
            )
            db.session.add(loc)
        db.session.commit()
        return jsonify({'ok': True, 'restored': Location.query.count()})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# health
@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'ok': True})

# simple location fetch by id
@app.route('/api/locations/<int:id>', methods=['GET'])
def get_location(id):
    loc = Location.query.get_or_404(id)
    return jsonify(loc.as_dict())

# --------------------
# Run
# --------------------
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050, debug=True)
